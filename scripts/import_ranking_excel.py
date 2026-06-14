from __future__ import annotations

import json
import re
from collections import OrderedDict
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_XLSX = Path(r"C:\Users\wesle\Downloads\Ranking Tennis 2026.xlsx")
PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = PROJECT_ROOT / "supabase" / "seeds" / "generated"
MONTH_SHEETS = [
    "JANEIRO",
    "FEVEREIRO",
    "MARCO",
    "ABRIL",
    "MAIO",
    "JUN",
    "JUL",
    "AGO",
    "SET",
    "OUT",
    "NOV",
    "DEZ",
]
RESULT_COLUMN_PAIRS = [(2, 3), (5, 6), (8, 9), (11, 12), (14, 15)]


@dataclass
class RankingRow:
    sheet_name: str
    season_year: int
    month_label: str
    player_name: str
    ranking_position: int | None
    points: float
    matches_played: float | None


@dataclass
class LegacyMatch:
    sheet_name: str
    season_year: int
    match_date: str
    team_a_names: list[str]
    team_b_names: list[str]
    team_a_scores: list[int]
    team_b_scores: list[int]
    result_summary: str
    source_cells: list[str]


def normalize_name(value: str) -> str:
    value = value.strip().upper()
    return re.sub(r"\s+", " ", value)


def normalize_token(value: str) -> str:
    replacements = str.maketrans({
        "Á": "A",
        "À": "A",
        "Â": "A",
        "Ã": "A",
        "É": "E",
        "Ê": "E",
        "Í": "I",
        "Ó": "O",
        "Ô": "O",
        "Õ": "O",
        "Ú": "U",
        "Ç": "C",
    })
    return normalize_name(value).translate(replacements)


def is_date_value(value: Any) -> bool:
    return isinstance(value, (datetime, date))


def safe_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def resolve_sheet_name(workbook, expected_name: str) -> str:
    normalized_expected = normalize_token(expected_name)
    for sheet_name in workbook.sheetnames:
        if normalize_token(sheet_name) == normalized_expected:
            return sheet_name
    raise KeyError(f"Aba nao encontrada para {expected_name}")


def extract_month_year(sheet_name: str, header_cells: list[Any]) -> int:
    for cell in header_cells:
        if isinstance(cell, datetime):
            return cell.year
    if sheet_name in {"JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"}:
        return 2025
    return 2026


def collect_ranking_rows(workbook) -> list[RankingRow]:
    rows: list[RankingRow] = []
    for month_key in MONTH_SHEETS:
        sheet_name = resolve_sheet_name(workbook, month_key)
        sheet = workbook[sheet_name]
        header_row = 4
        header_values = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
        season_year = extract_month_year(sheet_name, header_values)

        points_col = None
        games_col = None
        for idx, value in enumerate(header_values, start=1):
            if isinstance(value, str) and "PONTUA" in value.upper():
                points_col = idx
            if isinstance(value, str) and "JOGOS" in value.upper():
                games_col = idx

        for row in range(5, sheet.max_row + 1):
            raw_name = sheet.cell(row, 2).value if sheet_name != "FEVEREIRO" else sheet.cell(row, 3).value
            if not isinstance(raw_name, str):
                continue
            name = raw_name.strip()
            if not name or name.upper() == "RESULTADOS":
                continue
            if "/" in name or name.upper() in {"CHUVA", "#REF!", "RANKING"}:
                continue

            position_cell = sheet.cell(row, 1 if sheet_name != "FEVEREIRO" else 2).value
            ranking_position = int(position_cell) if isinstance(position_cell, int) else None
            points = safe_float(sheet.cell(row, points_col).value) if points_col else 0.0
            matches_played = safe_float(sheet.cell(row, games_col).value) if games_col else None

            if points == 0 and (matches_played or 0) == 0:
                continue

            rows.append(
                RankingRow(
                    sheet_name=sheet_name,
                    season_year=season_year,
                    month_label=sheet_name,
                    player_name=name,
                    ranking_position=ranking_position,
                    points=points,
                    matches_played=matches_played,
                )
            )
    return rows


def split_team_and_scores(text: str, reverse: bool = False) -> tuple[list[str], list[int]]:
    compact = re.sub(r"\s+", " ", text.strip())
    if reverse:
        match = re.match(r"^(?P<scores>[\d\(\)WOwo xX\s]+)\s+(?P<names>[A-Za-zÀ-ÿ/ ]+)$", compact)
        if not match:
            return [], []
        score_tokens = re.findall(r"\d+", match.group("scores"))
        names = [part.strip() for part in match.group("names").split("/") if part.strip()]
        return names, [int(token) for token in score_tokens]

    match = re.match(r"^(?P<names>[A-Za-zÀ-ÿ/ ]+?)\s+(?P<scores>[\d\(\)WOwo xX\s]+)$", compact)
    if not match:
        return [], []
    score_tokens = re.findall(r"\d+", match.group("scores"))
    names = [part.strip() for part in match.group("names").split("/") if part.strip()]
    return names, [int(token) for token in score_tokens]


def parse_match_row(left_text: str, right_text: str, match_date: datetime, sheet_name: str, source_cells: list[str]) -> LegacyMatch | None:
    if not left_text or not right_text:
        return None
    if "CHUVA" in left_text.upper() or "CHUVA" in right_text.upper():
        return None

    team_a_names, team_a_scores = split_team_and_scores(left_text, reverse=False)
    team_b_names, team_b_scores = split_team_and_scores(right_text, reverse=True)
    if len(team_a_names) != 2 or len(team_b_names) != 2:
        return None
    if not team_a_scores or len(team_a_scores) != len(team_b_scores):
        return None

    summary = " / ".join(f"{a}-{b}" for a, b in zip(team_a_scores, team_b_scores))
    return LegacyMatch(
        sheet_name=sheet_name,
        season_year=match_date.year,
        match_date=match_date.date().isoformat(),
        team_a_names=team_a_names,
        team_b_names=team_b_names,
        team_a_scores=team_a_scores,
        team_b_scores=team_b_scores,
        result_summary=summary,
        source_cells=source_cells,
    )


def collect_legacy_matches(workbook) -> list[LegacyMatch]:
    matches: list[LegacyMatch] = []
    for month_key in MONTH_SHEETS:
        sheet_name = resolve_sheet_name(workbook, month_key)
        sheet = workbook[sheet_name]
        result_anchor = None
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row, col).value
                if isinstance(value, str) and value.strip().upper() == "RESULTADOS":
                    result_anchor = row
                    break
            if result_anchor:
                break

        if not result_anchor:
            continue

        for row in range(result_anchor + 1, sheet.max_row + 1):
            for col_a, col_b in RESULT_COLUMN_PAIRS:
                date_value = sheet.cell(row, col_a).value
                if not is_date_value(date_value):
                    continue

                cursor = row + 1
                while cursor <= sheet.max_row:
                    left_value = sheet.cell(cursor, col_a).value
                    right_value = sheet.cell(cursor, col_b).value
                    if is_date_value(left_value):
                        break
                    if left_value is None and right_value is None:
                        cursor += 1
                        continue

                    if isinstance(left_value, str) and isinstance(right_value, str):
                        parsed = parse_match_row(
                            left_value,
                            right_value,
                            date_value,
                            sheet_name,
                            [f"{chr(64 + col_a)}{cursor}", f"{chr(64 + col_b)}{cursor}"],
                        )
                        if parsed:
                            matches.append(parsed)
                    cursor += 1
    return matches


def collect_players(ranking_rows: list[RankingRow], legacy_matches: list[LegacyMatch]) -> tuple[list[dict[str, Any]], list[str]]:
    ranked_names = OrderedDict((normalize_name(row.player_name), row.player_name) for row in ranking_rows)
    all_names = OrderedDict(ranked_names)
    unresolved_aliases: OrderedDict[str, None] = OrderedDict()

    for match in legacy_matches:
        for raw_name in [*match.team_a_names, *match.team_b_names]:
            normalized = normalize_name(raw_name)
            if normalized not in all_names:
                unresolved_aliases[raw_name] = None

    players = [
        {
            "display_name": source_name.title(),
            "full_name": source_name.title(),
            "normalized_name": normalized,
            "status": "active",
        }
        for normalized, source_name in all_names.items()
    ]
    return players, list(unresolved_aliases.keys())


def build_seed_payload(value_workbook, formula_workbook) -> dict[str, Any]:
    ranking_rows = collect_ranking_rows(value_workbook)
    legacy_matches = collect_legacy_matches(formula_workbook)
    players, unresolved_aliases = collect_players(ranking_rows, legacy_matches)

    annual_rows: list[dict[str, Any]] = []
    annual_sheet = value_workbook["RANKING ANUAL"]
    for row in range(5, annual_sheet.max_row + 1):
        name = annual_sheet.cell(row, 2).value
        points = annual_sheet.cell(row, 15).value
        games_year = annual_sheet.cell(row, 16).value
        if isinstance(name, str) and points not in (None, ""):
            annual_rows.append({
                "player_name": name.strip(),
                "points": safe_float(points),
                "matches_played": safe_float(games_year),
            })

    return {
        "source_file": str(DEFAULT_XLSX),
        "players": players,
        "monthly_rankings": [asdict(row) for row in ranking_rows],
        "annual_rankings": annual_rows,
        "legacy_matches": [asdict(match) for match in legacy_matches],
        "unresolved_aliases": unresolved_aliases,
        "notes": [
            "A planilha mistura dados de 2026 e 2025 nas abas de junho a dezembro.",
            "Os jogos historicos foram reconstruidos a partir do bloco RESULTADOS e exigem conciliacao manual para apelidos nao mapeados.",
            "Estatisticas de sets e games completas so sao confiaveis para partidas com placar totalmente parseado.",
        ],
    }


def main() -> None:
    value_workbook = load_workbook(DEFAULT_XLSX, data_only=True)
    formula_workbook = load_workbook(DEFAULT_XLSX, data_only=False)
    payload = build_seed_payload(value_workbook, formula_workbook)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_DIR / "legacy-ranking-2026.json"
    output_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Arquivo gerado: {output_file}")
    print(f"Jogadores: {len(payload['players'])}")
    print(f"Rankings mensais: {len(payload['monthly_rankings'])}")
    print(f"Partidas legadas parseadas: {len(payload['legacy_matches'])}")
    if payload["unresolved_aliases"]:
        print("Apelidos nao conciliados:")
        for alias in payload["unresolved_aliases"]:
            print(f" - {alias}")


if __name__ == "__main__":
    main()
