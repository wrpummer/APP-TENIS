from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_XLSX = Path(r"C:\Users\wesle\Downloads\Ranking Tennis 2026.xlsx")


def main() -> None:
    workbook = load_workbook(DEFAULT_XLSX, data_only=False)
    summary: dict[str, Any] = {
        "file": str(DEFAULT_XLSX),
        "sheets": workbook.sheetnames,
        "dimensions": {},
    }

    for name in workbook.sheetnames:
        sheet = workbook[name]
        summary["dimensions"][name] = {
            "rows": sheet.max_row,
            "cols": sheet.max_column,
        }

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
